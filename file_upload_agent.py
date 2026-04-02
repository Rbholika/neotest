"""
agents/file_upload_agent.py — File Ingestion & Vector Store Agent
==================================================================
LangGraph StateGraph nodes:
    validate_file → extract_text → pii_check
        → chunk_text → embed_and_store → done

Supported formats: PDF, DOCX, TXT, CSV, JSON, XLSX
All text is PII-masked before embedding.

Usage:
    from agents.file_upload_agent import FileUploadAgent
    agent = FileUploadAgent()
    result = agent.invoke({"file_path": "/path/to/doc.pdf",
                           "collection": "incidents"})
"""

import os
import json
import csv
from typing import Any, Dict, Optional, List, TypedDict

from langgraph.graph import StateGraph, END
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_community.vectorstores import Chroma

from agents.base_agent import BaseAgent, AgentResponse
from core.logger import get_logger
from core.security import guardrails
from config import CONFIG

logger = get_logger("agent.file_upload")


# ── LangGraph State ───────────────────────────────────────────────────────────

class FileUploadState(TypedDict):
    file_path:        str
    file_type:        str
    collection:       str
    raw_text:         str
    masked_text:      str
    chunks:           List[str]
    num_chunks:       int
    pii_found:        List[str]
    vector_store_dir: str
    error:            Optional[str]


# ── Agent ─────────────────────────────────────────────────────────────────────

class FileUploadAgent(BaseAgent):

    def __init__(self):
        super().__init__(name="file_upload")
        self.splitter = RecursiveCharacterTextSplitter(
            chunk_size=CONFIG.CHUNK_SIZE,
            chunk_overlap=CONFIG.CHUNK_OVERLAP,
            separators=["\n\n", "\n", " ", ""],
        )

    # ── LangGraph nodes ───────────────────────────────────────────────────────

    def _node_validate(self, state: FileUploadState) -> FileUploadState:
        path = state["file_path"]

        if not os.path.exists(path):
            state["error"] = f"File not found: {path}"
            return state

        ext = os.path.splitext(path)[1].lower().lstrip(".")
        if ext not in CONFIG.ALLOWED_FILE_TYPES:
            state["error"] = (
                f"File type '.{ext}' not allowed. "
                f"Allowed: {CONFIG.ALLOWED_FILE_TYPES}"
            )
            return state

        size_mb = os.path.getsize(path) / (1024 * 1024)
        if size_mb > CONFIG.MAX_FILE_SIZE_MB:
            state["error"] = (
                f"File too large ({size_mb:.1f} MB > {CONFIG.MAX_FILE_SIZE_MB} MB)"
            )
            return state

        state["file_type"] = ext
        logger.info(f"File validated: {path} ({ext}, {size_mb:.2f} MB)")
        return state

    def _node_extract(self, state: FileUploadState) -> FileUploadState:
        path = state["file_path"]
        ext  = state["file_type"]
        text = ""

        try:
            if ext == "pdf":
                from pdfminer.high_level import extract_text
                text = extract_text(path)

            elif ext == "docx":
                from docx import Document
                doc  = Document(path)
                text = "\n".join(p.text for p in doc.paragraphs)

            elif ext == "txt":
                with open(path, "r", encoding="utf-8", errors="ignore") as f:
                    text = f.read()

            elif ext == "csv":
                rows = []
                with open(path, "r", encoding="utf-8", errors="ignore") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        rows.append(", ".join(f"{k}: {v}" for k, v in row.items()))
                text = "\n".join(rows)

            elif ext == "json":
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                text = json.dumps(data, indent=2, ensure_ascii=False)

            elif ext == "xlsx":
                import openpyxl
                wb   = openpyxl.load_workbook(path, data_only=True)
                rows = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        rows.append(", ".join(str(c) for c in row if c is not None))
                text = "\n".join(rows)

            if not text.strip():
                state["error"] = "Could not extract any text from the file."
            else:
                state["raw_text"] = text
                logger.info(f"Extracted {len(text)} chars from {path}")

        except Exception as exc:
            state["error"] = f"Text extraction failed: {exc}"

        return state

    def _node_pii_check(self, state: FileUploadState) -> FileUploadState:
        text   = state.get("raw_text", "")
        result = guardrails.check_input(text)

        state["pii_found"]   = [f.pii_type for f in result.pii_findings]
        state["masked_text"] = result.sanitised_text or text

        if result.pii_findings:
            types = list(set(state["pii_found"]))
            logger.warning(
                f"PII detected and masked in uploaded file: {types}"
            )
        return state

    def _node_chunk(self, state: FileUploadState) -> FileUploadState:
        text   = state.get("masked_text") or state.get("raw_text", "")
        chunks = self.splitter.split_text(text)
        state["chunks"]     = chunks
        state["num_chunks"] = len(chunks)
        logger.info(f"Split into {len(chunks)} chunks")
        return state

    def _node_embed_store(self, state: FileUploadState) -> FileUploadState:
        chunks     = state.get("chunks", [])
        collection = state.get("collection", "default")
        persist_dir = os.path.join(CONFIG.CHROMA_DIR, collection)

        try:
            embeddings = self._get_embeddings()
            vs = Chroma.from_texts(
                texts=chunks,
                embedding=embeddings,
                persist_directory=persist_dir,
                collection_name=collection,
                metadatas=[
                    {
                        "source":     os.path.basename(state["file_path"]),
                        "chunk_idx":  i,
                        "file_type":  state["file_type"],
                        "collection": collection,
                    }
                    for i in range(len(chunks))
                ],
            )
            vs.persist()
            state["vector_store_dir"] = persist_dir
            logger.info(
                f"Stored {len(chunks)} chunks in Chroma "
                f"collection='{collection}' at {persist_dir}"
            )
        except Exception as exc:
            state["error"] = f"Embedding/storage failed: {exc}"

        return state

    # ── Routing ───────────────────────────────────────────────────────────────

    def _route_on_error(self, state: FileUploadState) -> str:
        return "error_end" if state.get("error") else "next"

    # ── Graph construction ────────────────────────────────────────────────────

    def _build_graph(self):
        g = StateGraph(FileUploadState)

        g.add_node("validate_file",   self._node_validate)
        g.add_node("extract_text",    self._node_extract)
        g.add_node("pii_check",       self._node_pii_check)
        g.add_node("chunk_text",      self._node_chunk)
        g.add_node("embed_and_store", self._node_embed_store)
        g.add_node("error_end",       lambda s: s)

        g.set_entry_point("validate_file")

        for src, nxt in [
            ("validate_file",   "extract_text"),
            ("extract_text",    "pii_check"),
            ("pii_check",       "chunk_text"),
            ("chunk_text",      "embed_and_store"),
        ]:
            g.add_conditional_edges(
                src,
                self._route_on_error,
                {"next": nxt, "error_end": "error_end"},
            )

        g.add_edge("embed_and_store", END)
        g.add_edge("error_end",       END)
        return g.compile()

    # ── Public API ────────────────────────────────────────────────────────────

    def invoke(self, inputs: Dict[str, Any]) -> AgentResponse:
        """
        inputs = {
            "file_path":  str,          # absolute or relative path
            "collection": str           # Chroma collection name (default: "default")
        }
        """
        resp = AgentResponse(agent_name=self.name)
        try:
            initial: FileUploadState = {
                "file_path":        inputs["file_path"],
                "file_type":        "",
                "collection":       inputs.get("collection", "default"),
                "raw_text":         "",
                "masked_text":      "",
                "chunks":           [],
                "num_chunks":       0,
                "pii_found":        [],
                "vector_store_dir": "",
                "error":            None,
            }
            final = self._graph.invoke(initial)

            if final.get("error"):
                return resp.fail(final["error"])

            resp.output   = {
                "chunks":           final["num_chunks"],
                "collection":       final["collection"],
                "vector_store_dir": final["vector_store_dir"],
                "pii_types_found":  list(set(final["pii_found"])),
            }
            resp.metadata = resp.output
            if final["pii_found"]:
                resp.warn(
                    f"PII detected and masked: {list(set(final['pii_found']))}"
                )
        except Exception as exc:
            return resp.fail(str(exc))
        return resp
